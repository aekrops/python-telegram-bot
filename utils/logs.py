from openpyxl import load_workbook
import pandas as pd


def write_logs_to_excel(time, user, user_id, text, answer):
    writer = pd.ExcelWriter('logs.xlsx', engine='openpyxl')
    wb = load_workbook('logs.xlsx')
    ws = wb["Sheet"]

    df_new_u = pd.DataFrame({"user": [user]})
    df_new_uid = pd.DataFrame({"user_id": [user_id]})
    df_new_time = pd.DataFrame({"time": [time]})
    df_new_ut = pd.DataFrame({"user_text": [text]})
    df_new_ba = pd.DataFrame({"bot_answer": [answer]})

    for index, row in df_new_u.iterrows():
        cell = "A%d" % (index + 2)
        ws[cell] = row[0]

    for index, row in df_new_uid.iterrows():
        cell = "B%d" % (index + 2)
        ws[cell] = row[0]

    for index, row in df_new_time.iterrows():
        cell = "C%d" % (index + 2)
        ws[cell] = row[0]

    for index, row in df_new_ut.iterrows():
        cell = "D%d" % (index + 2)
        ws[cell] = row[0]

    for index, row in df_new_ba.iterrows():
        cell = "E%d" % (index + 2)
        ws[cell] = row[0]

    wb.save("logs.xlsx")
    print("log saved...")


def log(message, answer):
    from datetime import datetime
    time = str(datetime.now())
    user = str(message.from_user.first_name)
    user_id = str(message.from_user.id)
    text = str(message.text)
    bot_answer = str(answer)
    write_logs_to_excel(time, user, user_id, text, bot_answer)
